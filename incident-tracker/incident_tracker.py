"""Writer for the damaged-fitting tracker workbook.

One row per damaged fitting (a report can list several).  If the workbook
already exists (HD's real tracker), the writer maps its actual header row
by fuzzy matching - same approach as the fault tracker - so the team's
column layout is preserved.  Sheets rebuilt on every run:

* ``IR ANNEX <MMM-YYYY>``  - the month's fittings + report references,
  the draft annex for the monthly IR (claim from ADA AGL inventory).
* ``QA REVIEW``            - one row per processed report with flags.
* ``SUMMARY``              - per-month counts and IR status.
"""

import copy
import datetime
import os
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from incident_parse import fitting_type

DATA_SHEET = "DAMAGED FITTINGS"
QA_SHEET = "QA REVIEW"
SUMMARY_SHEET = "SUMMARY"
ANNEX_PREFIX = "IR ANNEX"

# canonical field -> header aliases (lowercased, punctuation stripped).
# Covers both our default layout and HD's real DAMAGE_FITTING_LIST
# (SL NO. | Date | Circuit | Fitting No. | Location | Fault | Quantity
#  | Action | AED Cost).
HEADER_ALIASES = {
    "sl_no":        ["sl no", "sl", "s no", "sno", "serial"],
    "date":         ["date", "incident date", "date of incident"],
    "location":     ["location", "area", "site"],
    "fitting_type": ["fitting type", "type of fitting", "light type", "circuit"],
    "fitting_ref":  ["fitting ref", "fitting reference", "fitting id",
                     "asset", "fitting no", "light ref", "fitting"],
    "fault":        ["fault", "damage", "defect"],
    "quantity":     ["quantity", "qty"],
    "action":       ["action", "action taken", "rectification"],
    "cost":         ["aed cost", "unit cost", "cost"],
    "wo_number":    ["wo number", "workorder no", "work order", "wo", "wo no"],
    "report_id":    ["report id", "incident report no", "report no", "report ref"],
    "report_file":  ["report file", "incident report", "report filename",
                     "attachment", "file"],
    "reported_by":  ["reported by", "reporting person", "found by"],
    "prepared_by":  ["prepared by", "report prepared"],
    "signed":       ["signed", "signature", "sign"],
    "ir_no":        ["ir no", "ir ref", "ir number", "claim no"],
    "ir_status":    ["ir status", "status", "claim status"],
    "remarks":      ["remarks", "comment", "notes"],
}

# columns the workflow depends on: appended to an existing workbook's
# header row when its own layout lacks them (HD's list has no WO /
# report reference / IR columns - adding them is the tool's purpose).
ESSENTIAL_FIELDS = [("wo_number", "WO NUMBER"), ("report_id", "REPORT ID"),
                    ("ir_no", "IR NO"), ("ir_status", "IR STATUS")]

DEFAULT_HEADERS = [
    ("sl_no", "SL NO"), ("date", "DATE"), ("location", "LOCATION"),
    ("fitting_type", "FITTING TYPE"), ("fitting_ref", "FITTING REF"),
    ("wo_number", "WO NUMBER"), ("report_id", "REPORT ID"),
    ("report_file", "REPORT FILE"), ("reported_by", "REPORTED BY"),
    ("prepared_by", "PREPARED BY"), ("signed", "SIGNED"),
    ("ir_no", "IR NO"), ("ir_status", "IR STATUS"), ("remarks", "REMARKS"),
]

HEAD_FONT = Font(bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor="4472C4")
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _norm(s):
    return re.sub(r"[^a-z0-9 ]", "", str(s or "").lower()).strip()


def _month_key(d):
    return d.strftime("%b-%Y").upper() if d else "UNDATED"


def open_tracker(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = DATA_SHEET
    _write_default_header(ws)
    return wb


def _write_default_header(ws):
    for col, (_, title) in enumerate(DEFAULT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=title)
        c.font, c.fill, c.alignment = HEAD_FONT, HEAD_FILL, CENTER
        ws.column_dimensions[c.column_letter].width = max(12, len(title) + 4)
    ws.freeze_panes = "A2"


def find_data_sheet(wb):
    """Prefer an existing sheet whose name mentions damage/fitting; else
    the sheet whose header row maps the most known columns (HD's real
    list lives on a plain 'Sheet1'); else create our default sheet."""
    for name in wb.sheetnames:
        n = _norm(name)
        if ("damag" in n or "fitting" in n) and not n.startswith(_norm(ANNEX_PREFIX)):
            return wb[name]
    best, best_n = None, 0
    for name in wb.sheetnames:
        if name in (QA_SHEET, SUMMARY_SHEET) or name.startswith(ANNEX_PREFIX):
            continue
        _, cmap = _scan_headers(wb[name])
        if len(cmap) > best_n:
            best, best_n = wb[name], len(cmap)
    if best is not None and best_n >= 4:
        return best
    ws = wb.create_sheet(DATA_SHEET, 0)
    _write_default_header(ws)
    return ws


def _scan_headers(ws):
    """Find the row matching the most header aliases; no side effects."""
    best_row, best_map = None, {}
    for row in range(1, min(ws.max_row, 10) + 1):
        cells = {col: _norm(ws.cell(row=row, column=col).value)
                 for col in range(1, ws.max_column + 1)
                 if ws.cell(row=row, column=col).value not in (None, "")}
        taken, cmap = set(), {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in sorted(aliases, key=len, reverse=True):
                hit = next((c for c, v in cells.items()
                            if c not in taken and (v == alias or
                                                   (len(alias) > 3 and alias in v))), None)
                if hit:
                    cmap[field] = hit
                    taken.add(hit)
                    break
        if len(cmap) > len(best_map):
            best_row, best_map = row, cmap
    return best_row, best_map


def map_headers(ws):
    """Return (header_row, {field: column}) by fuzzy-matching the header
    row - works on HD's real workbook whatever the exact wording is.
    Aliases are tried longest-first and each column is claimed once.
    Falls back to writing our default header on an unrecognisable sheet.
    """
    best_row, best_map = _scan_headers(ws)
    if not best_map or len(best_map) < 3:
        # no recognisable header: write ours on row 1
        _write_default_header(ws)
        return 1, {f: i + 1 for i, (f, _) in enumerate(DEFAULT_HEADERS)}
    return best_row, best_map


def ensure_essential_columns(ws, header_row, cmap):
    """Append WO / report-reference / IR columns to an existing layout
    that lacks them, after its last used header column."""
    last = max([c for c in cmap.values()] +
               [col for col in range(1, ws.max_column + 1)
                if ws.cell(row=header_row, column=col).value not in (None, "")] or [0])
    style_from = ws.cell(row=header_row, column=last) if last else None
    for field, title in ESSENTIAL_FIELDS:
        if field in cmap:
            continue
        last += 1
        cell = ws.cell(row=header_row, column=last, value=title)
        if style_from is not None:
            cell.font = copy.copy(style_from.font)
            cell.fill = copy.copy(style_from.fill)
            cell.alignment = copy.copy(style_from.alignment)
        ws.column_dimensions[cell.column_letter].width = max(12, len(title) + 4)
        cmap[field] = last
    return cmap


def _existing_keys(ws, header_row, cmap):
    keys = {}
    ref_col, wo_col = cmap.get("fitting_ref"), cmap.get("wo_number")
    for row in range(header_row + 1, ws.max_row + 1):
        ref = str(ws.cell(row=row, column=ref_col).value or "").strip() if ref_col else ""
        wo = str(ws.cell(row=row, column=wo_col).value or "").strip() if wo_col else ""
        wo = re.sub(r"\.0$", "", wo)  # excel float artefacts
        if ref or wo:
            keys[(ref.upper(), wo)] = row
    return keys


def upsert_fittings(wb, rec, flags=None, catalog=None):
    """Add one row per fitting from a parsed record; update in place when
    the (fitting ref, WO) pair is already tracked.  Returns
    (added, updated, warnings)."""
    ws = find_data_sheet(wb)
    header_row, cmap = map_headers(ws)
    ensure_essential_columns(ws, header_row, cmap)
    keys = _existing_keys(ws, header_row, cmap)
    added = updated = 0
    warnings = []
    next_sl = 1
    if "sl_no" in cmap:
        existing_sl = [ws.cell(row=r, column=cmap["sl_no"]).value
                       for r in range(header_row + 1, ws.max_row + 1)]
        numeric = [int(v) for v in existing_sl
                   if isinstance(v, (int, float)) or (str(v or "").isdigit())]
        next_sl = (max(numeric) if numeric else 0) + 1

    for f in rec.get("fittings", []):
        ref = (f.get("ref") or "").upper()
        wo = str(f.get("wo") or rec.get("wo_number") or "")
        key = (ref, wo)
        # same WO already tracked under a different fitting ref -> warn
        for (kref, kwo), krow in keys.items():
            if wo and kwo == wo and kref != ref:
                warnings.append("WO %s already tracked for fitting %s (row %d)"
                                % (wo, kref or "?", krow))
        if key in keys:
            row, updated = keys[key], updated + 1
        elif (ref, "") in keys:
            # fitting already listed without a WO (e.g. HD's historical
            # rows before this tool): enrich that row instead of duplicating
            row, updated = keys.pop((ref, "")), updated + 1
            keys[key] = row
        else:
            row, added = ws.max_row + 1, added + 1
            keys[key] = row
            if "sl_no" in cmap:
                ws.cell(row=row, column=cmap["sl_no"], value=next_sl)
                next_sl += 1
        ftype = fitting_type(ref)
        cat = (catalog or {}).get(ftype, {})
        resolution = (rec.get("resolution") or "").strip().split("\n")[0]
        values = {
            "date": rec.get("incident_date") or rec.get("report_date"),
            "location": rec.get("location"),
            "fitting_type": ftype,
            "fitting_ref": ref,
            "fault": "Damage fitting",
            "quantity": 1,
            "action": resolution or "Damage fitting replaced",
            "cost": cat.get("unit_price"),
            "wo_number": wo,
            "report_id": "%s-%s" % (rec.get("report_id"), rec.get("report_serial"))
                         if rec.get("report_serial") else rec.get("report_id"),
            "report_file": rec.get("file"),
            "reported_by": rec.get("reporting_person"),
            "prepared_by": rec.get("prepared_by"),
            "signed": "YES" if rec.get("sig_prepared") else "NO",
        }
        for field, val in values.items():
            col = cmap.get(field)
            if col is None or val in (None, ""):
                continue
            cell = ws.cell(row=row, column=col)
            cell.value = val
            if isinstance(val, datetime.date):
                cell.number_format = "DD-MMM-YY"
        if flags and "remarks" in cmap:
            cell = ws.cell(row=row, column=cmap["remarks"])
            existing = str(cell.value or "")
            note = "CHECK: " + "; ".join(flags)
            if note not in existing:
                cell.value = (existing + " | " if existing else "") + note
    return added, updated, warnings


def _data_rows(ws, header_row, cmap):
    rows = []
    for row in range(header_row + 1, ws.max_row + 1):
        def get(field):
            col = cmap.get(field)
            return ws.cell(row=row, column=col).value if col else None
        if not (get("fitting_ref") or get("wo_number")):
            continue
        d = get("date")
        if isinstance(d, datetime.datetime):
            d = d.date()
        rows.append({f: get(f) for f in HEADER_ALIASES} | {"date": d, "_row": row})
    return rows


def rebuild_annex_sheets(wb, recent_months=6):
    """Rebuild one 'IR ANNEX <MMM-YYYY>' sheet per month present in the
    data - the draft annex for the monthly IR raised from HD's tracker.
    Only the most recent `recent_months` months get a sheet (0 = all), so
    a workbook with years of history doesn't grow dozens of annex tabs."""
    ws = find_data_sheet(wb)
    header_row, cmap = map_headers(ws)
    rows = _data_rows(ws, header_row, cmap)

    for name in [n for n in wb.sheetnames if n.startswith(ANNEX_PREFIX)]:
        del wb[name]

    by_month = {}
    for r in rows:
        by_month.setdefault(_month_key(r["date"]), []).append(r)
    if recent_months and len(by_month) > recent_months:
        dated = sorted((m for m in by_month if m != "UNDATED"),
                       key=lambda m: datetime.datetime.strptime(m, "%b-%Y"))
        keep = set(dated[-recent_months:]) | ({"UNDATED"} & set(by_month))
        by_month = {m: v for m, v in by_month.items() if m in keep}

    cols = [("SL NO", 8), ("DATE", 12), ("LOCATION", 24), ("FITTING TYPE", 14),
            ("FITTING REF", 18), ("WO NUMBER", 14), ("INCIDENT REPORT REF", 26),
            ("IR NO", 12), ("IR STATUS", 12)]
    for month in sorted(by_month, key=lambda m: (m == "UNDATED",
                                                 datetime.datetime.strptime(m, "%b-%Y")
                                                 if m != "UNDATED" else datetime.datetime.max)):
        sheet = wb.create_sheet("%s %s" % (ANNEX_PREFIX, month))
        title = sheet.cell(row=1, column=1,
                           value="IR ANNEX - AGL FITTINGS DAMAGED BY THIRD PARTIES - %s" % month)
        title.font = Font(bold=True, size=12)
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(cols))
        for c, (t, w) in enumerate(cols, start=1):
            cell = sheet.cell(row=2, column=c, value=t)
            cell.font, cell.fill, cell.alignment = HEAD_FONT, HEAD_FILL, CENTER
            sheet.column_dimensions[cell.column_letter].width = w
        for i, r in enumerate(sorted(by_month[month],
                                     key=lambda x: (x["date"] or datetime.date.max,
                                                    str(x["fitting_ref"]))), start=1):
            vals = [i, r["date"], r["location"], r["fitting_type"], r["fitting_ref"],
                    r["wo_number"], r["report_id"] or r["report_file"],
                    r["ir_no"], r["ir_status"]]
            for c, v in enumerate(vals, start=1):
                cell = sheet.cell(row=2 + i, column=c, value=v)
                if isinstance(v, datetime.date):
                    cell.number_format = "DD-MMM-YY"
        n = len(by_month[month])
        total = sheet.cell(row=3 + n + 1, column=1,
                           value="TOTAL FITTINGS: %d" % n)
        total.font = Font(bold=True)


def rebuild_qa_sheet(wb, reviews):
    """reviews: list of (record, flags) processed so far (from the ledger +
    this run).  One row per report file, flagged rows highlighted."""
    if QA_SHEET in wb.sheetnames:
        del wb[QA_SHEET]
    ws = wb.create_sheet(QA_SHEET)
    cols = [("REPORT FILE", 52), ("REPORT DATE", 12), ("WO", 12),
            ("FITTINGS", 22), ("PREPARER SIGNED", 16), ("ADA SIGNED", 12),
            ("ANNEX PHOTOS", 13), ("RESULT", 10), ("FLAGS", 70)]
    for c, (t, w) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=t)
        cell.font, cell.fill, cell.alignment = HEAD_FONT, HEAD_FILL, CENTER
        ws.column_dimensions[cell.column_letter].width = w
    ws.freeze_panes = "A2"
    for i, (rec, flags) in enumerate(sorted(reviews,
                                            key=lambda x: str(x[0].get("file"))), start=2):
        vals = [rec.get("file"), rec.get("report_date"), rec.get("wo_number"),
                ", ".join(f["ref"] for f in rec.get("fittings", [])),
                "YES" if rec.get("sig_prepared") else "NO",
                "YES" if rec.get("sig_submitted") else "NO",
                rec.get("annex_photos"),
                "CHECK" if flags else "OK",
                "; ".join(flags)]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            if isinstance(v, datetime.date):
                cell.number_format = "DD-MMM-YY"
        if flags:
            for c in range(1, len(cols) + 1):
                ws.cell(row=i, column=c).fill = WARN_FILL


def rebuild_summary_sheet(wb):
    ws_data = find_data_sheet(wb)
    header_row, cmap = map_headers(ws_data)
    rows = _data_rows(ws_data, header_row, cmap)
    if SUMMARY_SHEET in wb.sheetnames:
        del wb[SUMMARY_SHEET]
    ws = wb.create_sheet(SUMMARY_SHEET, 1)
    cols = [("MONTH", 12), ("FITTINGS DAMAGED", 18), ("REPORTS", 10),
            ("IR RAISED", 10), ("IR PENDING", 11)]
    for c, (t, w) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=c, value=t)
        cell.font, cell.fill, cell.alignment = HEAD_FONT, HEAD_FILL, CENTER
        ws.column_dimensions[cell.column_letter].width = w
    by_month = {}
    for r in rows:
        by_month.setdefault(_month_key(r["date"]), []).append(r)
    for i, month in enumerate(sorted(by_month,
                                     key=lambda m: (m == "UNDATED",
                                                    datetime.datetime.strptime(m, "%b-%Y")
                                                    if m != "UNDATED" else datetime.datetime.max)),
                              start=2):
        items = by_month[month]
        reports = {str(r["report_id"] or r["report_file"]) for r in items}
        raised = sum(1 for r in items if r["ir_no"])
        ws.cell(row=i, column=1, value=month)
        ws.cell(row=i, column=2, value=len(items))
        ws.cell(row=i, column=3, value=len(reports))
        ws.cell(row=i, column=4, value=raised)
        pend = ws.cell(row=i, column=5, value=len(items) - raised)
        if len(items) - raised:
            pend.fill = WARN_FILL


def save_tracker(wb, path):
    wb.save(path)
