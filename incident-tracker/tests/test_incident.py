"""Tests for the incident report tracker (fabricated PDFs, fictional data)."""

import os
import sys

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from incident_parse import parse_incident_pdf, review_record, fitting_type
from incident_tracker import (open_tracker, upsert_fittings, rebuild_annex_sheets,
                              rebuild_qa_sheet, rebuild_summary_sheet, save_tracker,
                              find_data_sheet, map_headers)
from tests.make_test_pdfs import make_all, make_report


@pytest.fixture(scope="module")
def pdfs(tmp_path_factory):
    outdir = tmp_path_factory.mktemp("pdfs")
    return make_all(str(outdir))


def test_parse_good_report(pdfs):
    rec = parse_incident_pdf(pdfs[0])
    assert rec["report_id"] == "20260114"
    assert rec["report_serial"] == "01"
    assert rec["wo_number"] == "43900001"
    assert str(rec["report_date"]) == "2026-01-14"
    assert rec["reporting_person"] == "Testperson"
    assert rec["reporting_manager"] == "Testmanager Lead"
    assert rec["location"] == "Twy C b/w C1 and C2"
    assert rec["area_of_incident"] == "Twy C b/w C1 and C2"
    assert (rec["time_occurred"], rec["time_attended"], rec["time_rectified"]) == \
        ("08:00", "08:10", "09:00")
    assert rec["fittings"] == [{"ref": "TEC102-05/001", "wo": "43900001"}]
    assert rec["prepared_by"] == "Test Preparer"
    assert rec["submitted_to"] == "Test Receiver"
    assert rec["sig_prepared"] is True
    assert rec["sig_submitted"] is False
    assert rec["annex_photos"] == 2
    assert rec["annex_refs"] == ["TEC102-05/001"]
    assert "Airside" in rec["categories"]
    assert review_record(rec) == []


def test_parse_multi_fitting(pdfs):
    rec = parse_incident_pdf(pdfs[1])
    assert rec["fittings"] == [{"ref": "TEC102-07/010", "wo": "43900002"},
                               {"ref": "TEC102-07/011", "wo": "43900003"}]


def test_bad_report_flags(pdfs):
    rec = parse_incident_pdf(pdfs[2])
    flags = review_record(rec)
    joined = " ; ".join(flags)
    assert "signature image missing" in joined
    assert "!= Area of Incident" in joined
    assert "days after incident" in joined
    assert "outside 100-800" in joined
    assert "no annex photos" in joined


def test_fitting_type():
    assert fitting_type("TEC102-01/015") == "TEC"
    assert fitting_type("LICATC-36/024") == "LICATC"


def test_tracker_upsert_and_rebuild(pdfs, tmp_path):
    path = str(tmp_path / "tracker.xlsx")
    wb = open_tracker(path)
    recs = [parse_incident_pdf(p) for p in pdfs]
    total_added = 0
    for rec in recs:
        added, updated, _ = upsert_fittings(wb, rec, review_record(rec))
        total_added += added
    assert total_added == 4  # 1 + 2 + 1 fittings

    # re-upsert the same records: rows update in place, no duplicates
    for rec in recs:
        added, updated, _ = upsert_fittings(wb, rec)
        assert added == 0 and updated == len(rec["fittings"])

    rebuild_annex_sheets(wb)
    rebuild_qa_sheet(wb, [(r, review_record(r)) for r in recs])
    rebuild_summary_sheet(wb)
    save_tracker(wb, path)

    wb2 = load_workbook(path)
    assert "IR ANNEX JAN-2026" in wb2.sheetnames
    assert "IR ANNEX FEB-2026" in wb2.sheetnames
    jan = wb2["IR ANNEX JAN-2026"]
    refs = [row[4] for row in jan.iter_rows(min_row=3, values_only=True) if row[4]]
    assert refs == ["TEC102-05/001", "TEC102-07/010", "TEC102-07/011"]
    qa = wb2["QA REVIEW"]
    results = [row[7] for row in qa.iter_rows(min_row=2, values_only=True) if row[0]]
    assert results.count("OK") == 2 and results.count("CHECK") == 1


def test_duplicate_wo_warning(pdfs, tmp_path):
    wb = open_tracker(str(tmp_path / "t.xlsx"))
    rec = parse_incident_pdf(pdfs[0])
    upsert_fittings(wb, rec)
    other = dict(rec, fittings=[{"ref": "TEC102-99/099", "wo": rec["wo_number"]}])
    _, _, warns = upsert_fittings(wb, other)
    assert warns and "already tracked" in warns[0]


def test_header_automap_hd_style(pdfs, tmp_path):
    """The writer must adapt to HD's own workbook layout, not impose ours."""
    path = str(tmp_path / "hd.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Damaged Fittings 2026"
    ws.append(["S.No", "Date", "Area", "Fitting ID", "Work Order",
               "Incident Report No", "Status", "Remarks"])
    ws.append([1, "01-Jan-26", "Twy Z", "TEC999-01/001", "43800000",
               "20260101-01", "IR RAISED", ""])
    wb.save(path)

    wb = open_tracker(path)
    ws = find_data_sheet(wb)
    assert ws.title == "Damaged Fittings 2026"
    header_row, cmap = map_headers(ws)
    assert header_row == 1
    assert cmap["fitting_ref"] is not None
    assert cmap["wo_number"] is not None

    rec = parse_incident_pdf(pdfs[0])
    added, _, _ = upsert_fittings(wb, rec)
    assert added == 1
    row = [c.value for c in ws[3]]
    assert "TEC102-05/001" in row and "43900001" in [str(v) for v in row]
    # existing row untouched
    assert ws.cell(row=2, column=4).value == "TEC999-01/001"


def test_run_incident_end_to_end(pdfs, tmp_path, monkeypatch):
    """Manual-mode run twice: second run is a no-op (ledger idempotency)."""
    import shutil
    import run_incident

    workdir = tmp_path / "suite"
    (workdir / "incident_inbox").mkdir(parents=True)
    for p in pdfs:
        shutil.copy(p, workdir / "incident_inbox" / os.path.basename(p))
    monkeypatch.setattr(run_incident, "HERE", str(workdir))

    assert run_incident.main(["--manual"]) == 0
    wb = load_workbook(workdir / "DAMAGED_FITTING_TRACKER.xlsx")
    data = wb[wb.sheetnames[0]]
    n_rows = sum(1 for r in data.iter_rows(min_row=2, values_only=True) if r[4])
    assert n_rows == 4

    assert run_incident.main(["--manual"]) == 0
    wb = load_workbook(workdir / "DAMAGED_FITTING_TRACKER.xlsx")
    data = wb[wb.sheetnames[0]]
    n_rows2 = sum(1 for r in data.iter_rows(min_row=2, values_only=True) if r[4])
    assert n_rows2 == n_rows
